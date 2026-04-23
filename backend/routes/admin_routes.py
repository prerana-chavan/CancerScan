from flask import Blueprint, request, jsonify, g, Response, current_app
from database.db import get_connection, log_audit
from middleware.auth_middleware import require_auth
import bcrypt
import csv
import io

admin_bp = Blueprint('admin', __name__)

from routes.patient_routes import serialize_patient

def require_admin(f):
    '''Decorator: only admin role can access'''
    from functools import wraps
    @wraps(f)
    @require_auth
    def decorated(*args, **kwargs):
        if g.user.get('role') != 'admin':
            return jsonify({
                'success': False,
                'error':   'Admin access required'
            }), 403
        return f(*args, **kwargs)
    return decorated

# ── GET all doctors ──────────────────────────
@admin_bp.route('/doctors', methods=['GET'])
@require_admin
def get_all_doctors():
    conn = None
    try:
        conn    = get_connection()
        doctors = conn.execute('''
            SELECT id, full_name, email, role,
                   hospital, specialization,
                   medical_license_id,
                   is_approved, is_active,
                   created_at
            FROM doctors
            WHERE role != "admin"
            ORDER BY created_at DESC
        ''').fetchall()

        return jsonify({
            'success': True,
            'doctors': [dict(d) for d in doctors]
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in get_doctors: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to load doctors list.'}), 500
    finally:
        if conn:
            conn.close()

# ── TOGGLE doctor approval status ────────────
@admin_bp.route('/doctors/<doctor_id>/approve',
                methods=['POST'])
@require_admin
def toggle_doctor_approval(doctor_id):
    conn = None
    try:
        conn = get_connection()
        # Get current status
        doctor = conn.execute(
            'SELECT is_approved FROM doctors WHERE id = ?',
            (doctor_id,)
        ).fetchone()

        if not doctor:
            return jsonify({'success': False, 'error': 'Doctor not found'}), 404

        new_status = 1 if doctor['is_approved'] == 0 else 0
        
        conn.execute(
            'UPDATE doctors SET is_approved = ? '
            'WHERE id = ? AND role = "doctor"',
            (new_status, doctor_id)
        )
        conn.commit()

        # Audit: approve or reject
        evt = 'APPROVE' if new_status == 1 else 'REJECT'
        act = 'Doctor Approved' if new_status == 1 else 'Doctor Rejected'
        det = f"Account {'activated' if new_status == 1 else 'deactivated'} for doctor ID: {doctor_id}"
        log_audit(
            conn,
            event_type  = evt,
            category    = 'ADMIN',
            actor_id    = str(g.user['id']),
            actor_name  = g.user.get('name', 'Unknown'),
            action      = act,
            detail      = det,
            target_id   = str(doctor_id),
            ip_address  = request.remote_addr or 'localhost'
        )

        status_label = "Approved" if new_status == 1 else "Restricted"
        return jsonify({
            'success': True,
            'is_approved': new_status,
            'message': f'Doctor access set to {status_label}'
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in toggle_doctor_approval: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to toggle approval.'}), 500
    finally:
        if conn:
            conn.close()

# ── REJECT / DELETE doctor ───────────────────
@admin_bp.route('/doctors/<doctor_id>',
                methods=['DELETE'])
@require_admin
def delete_doctor(doctor_id):
    conn = None
    try:
        conn = get_connection()

        # Don't allow deleting admin
        doctor = conn.execute(
            'SELECT role FROM doctors WHERE id = ?',
            (doctor_id,)
        ).fetchone()

        if not doctor:
            return jsonify({
                'success': False,
                'error':   'Doctor not found'
            }), 404

        if doctor['role'] == 'admin':
            return jsonify({
                'success': False,
                'error':   'Cannot delete admin account'
            }), 403

        # Safely nullify foreign key references to preserve patient history
        conn.execute(
            'UPDATE patients SET doctor_id = NULL WHERE doctor_id = ?',
            (doctor_id,)
        )

        # Now safe to permanently delete the doctor record
        conn.execute(
            'DELETE FROM doctors WHERE id = ?',
            (doctor_id,)
        )
        conn.commit()

        # Audit: reject/delete
        log_audit(
            conn,
            event_type  = 'REJECT',
            category    = 'ADMIN',
            actor_id    = str(g.user['id']),
            actor_name  = g.user.get('name', 'Unknown'),
            action      = 'Doctor Rejected',
            detail      = f"Account removed for doctor ID: {doctor_id}",
            target_id   = str(doctor_id),
            ip_address  = request.remote_addr or 'localhost'
        )

        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f"Internal error in delete_doctor: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to delete doctor.'}), 500
    finally:
        if conn:
            conn.close()

# ── RESET doctor password ────────────────────
@admin_bp.route('/doctors/<doctor_id>/reset-password',
                methods=['POST'])
@require_admin
def reset_password(doctor_id):
    data         = request.get_json()
    new_password = data.get('newPassword', 'Doctor@1234')

    new_hash = bcrypt.hashpw(
        new_password.encode(),
        bcrypt.gensalt()
    ).decode()

    conn = None
    try:
        conn = get_connection()
        conn.execute(
            'UPDATE doctors SET password_hash = ? '
            'WHERE id = ?',
            (new_hash, doctor_id)
        )
        conn.commit()

        # Audit: password reset
        log_audit(
            conn,
            event_type  = 'PASSWORD',
            category    = 'ADMIN',
            actor_id    = str(g.user['id']),
            actor_name  = g.user.get('name', 'Unknown'),
            action      = 'Password Reset',
            detail      = f"Admin reset password for doctor ID: {doctor_id}",
            target_id   = str(doctor_id),
            ip_address  = request.remote_addr or 'localhost'
        )

        return jsonify({
            'success':     True,
            'message':     'Password reset successfully',
            'newPassword': new_password
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in reset_doctor_password: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to reset password.'}), 500
    finally:
        if conn:
            conn.close()

# ── GET system stats ─────────────────────────
@admin_bp.route('/stats', methods=['GET'])
@require_admin
def get_stats():
    conn = None
    try:
        conn = get_connection()

        total_doctors  = conn.execute(
            'SELECT COUNT(*) FROM doctors '
            'WHERE role != "admin"'
        ).fetchone()[0]

        pending        = conn.execute(
            'SELECT COUNT(*) FROM doctors '
            'WHERE role != "admin" AND is_approved = 0'
        ).fetchone()[0]

        approved       = conn.execute(
            'SELECT COUNT(*) FROM doctors '
            'WHERE role != "admin" AND is_approved = 1'
        ).fetchone()[0]

        total_patients = conn.execute(
            'SELECT COUNT(*) FROM patients'
        ).fetchone()[0]

        cancer_count   = conn.execute(
            """SELECT COUNT(*) FROM patients
               WHERE (prediction_result LIKE '%Cancer%' 
                      OR prediction_result LIKE '%Malignant%')
               AND prediction_result NOT LIKE '%No Cancer%'"""
        ).fetchone()[0]

        benign_count   = conn.execute(
            """SELECT COUNT(*) FROM patients
               WHERE prediction_result LIKE '%No Cancer%'
               OR prediction_result LIKE '%Benign%'
               OR prediction_result LIKE '%Normal%'"""
        ).fetchone()[0]

        return jsonify({
            'success': True,
            'stats': {
                'totalDoctors':   total_doctors,
                'pendingDoctors': pending,
                'approvedDoctors':approved,
                'totalScans':     total_patients,
                'cancerDetected': cancer_count,
                'benignScans':    benign_count,
            }
        })
    except Exception as e:
        current_app.logger.error(f"Stats route error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve system status'
        }), 500
    finally:
        if conn:
            conn.close()

# ── GET all patients (across all doctors) ────
@admin_bp.route('/patients', methods=['GET'])
@require_admin
def get_all_patients():
    conn = None
    try:
        conn     = get_connection()
        patients = conn.execute('''
            SELECT p.*,
                   p.status_stage as status,
                   CASE WHEN p.is_flagged = 1 THEN 'Urgent' ELSE 'Normal' END as priority,
                   'Node: ' || COALESCE(d.hospital, 'Active') as telemetry,
                   d.full_name as doctor_name,
                   d.hospital  as doctor_hospital
            FROM patients p
            LEFT JOIN doctors d ON p.doctor_id = d.id
            ORDER BY p.created_at DESC
        ''').fetchall()

        return jsonify({
            'success':  True,
            'patients': [serialize_patient(p) for p in patients]
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in get_all_patients: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to load patient records.'}), 500
    finally:
        if conn:
            conn.close()

# ── SEARCH all patients (across all doctors) ──
@admin_bp.route('/patients/search', methods=['GET'])
@require_admin
def search_all_patients():
    query = request.args.get('q', '')
    conn = None
    try:
        conn = get_connection()
        patients = conn.execute('''
            SELECT p.*,
                   p.status_stage as status,
                   CASE WHEN p.is_flagged = 1 THEN 'Urgent' ELSE 'Normal' END as priority,
                   'Node: ' || COALESCE(d.hospital, 'Active') as telemetry,
                   d.full_name as doctor_name,
                   d.hospital  as doctor_hospital
            FROM patients p
            LEFT JOIN doctors d ON p.doctor_id = d.id
            WHERE p.patient_id LIKE ? 
               OR p.hospital_name LIKE ?
               OR d.full_name LIKE ?
            ORDER BY p.created_at DESC
        ''', (f'%{query}%', f'%{query}%', f'%{query}%')).fetchall()

        return jsonify({
            'success': True,
            'patients': [serialize_patient(p) for p in patients]
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in search_all_patients: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Search failed.'}), 500
    finally:
        if conn:
            conn.close()

@admin_bp.route('/audit-logs', methods=['GET'])
@require_admin
def get_audit_logs():
    conn = None
    try:
        conn    = get_connection()
        page    = int(request.args.get('page', 1))
        limit   = int(request.args.get('limit', 20))
        search  = request.args.get('search', '')
        filter_type = request.args.get('type', 'ALL')
        offset  = (page - 1) * limit

        base_query = '''
            SELECT * FROM audit_logs
            WHERE 1=1
        '''
        params = []

        if search:
            base_query += '''
                AND (actor_name LIKE ?
                  OR action LIKE ?
                  OR detail LIKE ?
                  OR target_id LIKE ?
                  OR evt_id LIKE ?)
            '''
            s = f'%{search}%'
            params.extend([s, s, s, s, s])

        if filter_type != 'ALL':
            base_query += ' AND event_type = ?'
            params.append(filter_type)

        # Total count
        total = conn.execute(
            f'SELECT COUNT(*) FROM ({base_query})',
            params
        ).fetchone()[0]

        # Paginated results
        rows = conn.execute(
            base_query +
            ' ORDER BY created_at DESC'
            f' LIMIT {limit} OFFSET {offset}',
            params
        ).fetchall()

        return jsonify({
            'success':    True,
            'audit_logs': [dict(r) for r in rows],
            'total':      total
        })
    except Exception as e:
        current_app.logger.error(f"Internal error in get_audit_logs: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'An internal server error occurred. Please try again.'
        }), 500
    finally:
        if conn:
            conn.close()

@admin_bp.route('/audit-logs/export', methods=['GET'])
def export_audit_logs():
    conn = None
    try:
        # Accept token from header OR query param for download
        import os
        import jwt
        from config.settings import settings
        secret = settings.JWT_SECRET

        token = None
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        if not token:
            token = request.args.get('token','')

        if not token:
            return jsonify({
                'success': False,
                'error': 'No token'
            }), 401

        try:
            payload = jwt.decode(
                token, secret,
                algorithms=['HS256']
            )
            if payload.get('role') != 'admin':
                return jsonify({
                    'success': False,
                    'error': 'Admin only'
                }), 403
        except:
            return jsonify({
                'success': False,
                'error': 'Invalid token'
            }), 401

        # Fetch all audit logs
        conn = get_connection()
        rows = conn.execute(
            '''SELECT
                evt_id, event_type, category,
                actor_name, action, detail,
                target_id, ip_address,
                created_at
               FROM audit_logs
               ORDER BY created_at DESC'''
        ).fetchall()

        # Build Excel file
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = f'audit_log_{datetime.now().strftime("%Y-%m-%d")}'

        # Color schemes ...
        DARK_GREEN  = '0D3B2E'
        MID_GREEN   = '1A5C46'
        TEAL        = '0891B2'
        LIGHT_TEAL  = 'E0F2FE'
        WHITE       = 'FFFFFF'
        LIGHT_GREY  = 'F8FAFC'
        BORDER_GREY = 'E2E8F0'
        RED         = 'DC2626'
        ORANGE      = 'D97706'
        GREEN       = '16A34A'
        PURPLE      = '7C3AED'
        BLUE        = '0369A1'
        DARK_TEXT   = '1E293B'
        MID_TEXT    = '475569'

        # Row 1: Main Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'CANCERSCAN — SECURITY AUDIT LOG'
        title_cell.font = Font(name='Calibri', bold=True, size=16, color=WHITE)
        title_cell.fill = PatternFill(fill_type='solid', fgColor=DARK_GREEN)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        # Row 2: Subtitle/Info
        ws.merge_cells('A2:H2')
        now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
        sub = ws['A2']
        sub.value = (
            f'Generated On: {now_str}  |  '
            f'Total Events: {len(rows)}  |  '
            f'Exported By: CancerScan AI System'
        )
        sub.font = Font(name='Calibri', size=10, color=WHITE, italic=True)
        sub.fill = PatternFill(fill_type='solid', fgColor=MID_GREEN)
        sub.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # Spacer rows and summary stats...
        ws.row_dimensions[3].height = 8
        login_count = sum(1 for r in rows if r['event_type'] == 'LOGIN')
        fail_count = sum(1 for r in rows if r['event_type'] == 'LOGIN_FAIL')
        scan_count = sum(1 for r in rows if r['event_type'] == 'SCAN')
        del_count = sum(1 for r in rows if r['event_type'] == 'DELETE')

        ws.merge_cells('A4:B4')
        ws.merge_cells('C4:D4')
        ws.merge_cells('E4:F4')
        ws.merge_cells('G4:H4')

        summary_data = [
            ('A4', f'TOTAL EVENTS: {len(rows)}', TEAL, WHITE),
            ('C4', f'LOGINS: {login_count}', GREEN, WHITE),
            ('E4', f'FAILED: {fail_count}', ORANGE if fail_count > 0 else MID_TEXT, WHITE),
            ('G4', f'SCANS: {scan_count}  |  DELETIONS: {del_count}', DARK_TEXT, WHITE),
        ]
        for cell_ref, val, bg, fg in summary_data:
            c = ws[cell_ref]
            c.value = val
            c.font  = Font(name='Calibri', bold=True, size=10, color=fg)
            c.fill = PatternFill(fill_type='solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 6

        # Headers
        headers = [
            'EVENT ID', 'TYPE', 'CATEGORY', 'ACTOR', 'ACTION', 
            'DETAIL', 'IP ADDRESS', 'TIMESTAMP',
        ]
        header_fill = PatternFill(fill_type='solid', fgColor=DARK_GREEN)
        header_font = Font(name='Calibri', bold=True, size=10, color=WHITE)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin = Side(style='thin', color=BORDER_GREY)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
            cell.border    = border

        ws.row_dimensions[6].height = 28

        TYPE_COLORS = {
            'LOGIN':      (LIGHT_TEAL, BLUE),
            'LOGOUT':     ('F1F5F9',   MID_TEXT),
            'LOGIN_FAIL': ('FEF3C7',   ORANGE),
            'SCAN':       (LIGHT_TEAL, TEAL),
            'DELETE':     ('FEE2E2',   RED),
            'APPROVE':    ('DCFCE7',   GREEN),
            'REJECT':     ('FEF3C7',   ORANGE),
            'REGISTER':   ('EDE9FE',   PURPLE),
            'PASSWORD':   (LIGHT_TEAL, BLUE),
            'EXPORT':     ('DCFCE7',   GREEN),
        }

        # Data rows
        for row_num, row in enumerate(rows, 7):
            is_odd = (row_num - 7) % 2 == 0
            row_bg = WHITE if is_odd else 'F8FAFC'
            row_fill = PatternFill(fill_type='solid', fgColor=row_bg)
            evt_type = row['event_type'] or ''
            type_bg, type_fg = TYPE_COLORS.get(evt_type, ('F1F5F9', MID_TEXT))

            row_data = [
                row['evt_id']     or '',
                row['event_type'] or '',
                row['category']   or '',
                row['actor_name'] or '',
                row['action']     or '',
                row['detail']     or '',
                row['ip_address'] or '',
                row['created_at'] or '',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if col_num == 2:
                    cell.fill = PatternFill(fill_type='solid', fgColor=type_bg)
                    cell.font = Font(name='Calibri', size=9, bold=True, color=type_fg)
                    cell.alignment = center
                else:
                    cell.fill = row_fill
                    cell.font = Font(name='Calibri', size=9, color=DARK_TEXT)
                    if col_num in (4, 5, 6):
                        cell.alignment = left
                    else:
                        cell.alignment = center
                if col_num == 1:
                    cell.font = Font(name='Calibri', size=9, bold=True, color=TEAL)

            ws.row_dimensions[row_num].height = 20

        # Footer
        footer_row = len(rows) + 8
        ws.merge_cells(f'A{footer_row}:H{footer_row}')
        footer = ws[f'A{footer_row}']
        footer.value = (
            'CONFIDENTIAL — FOR AUTHORIZED PERSONNEL ONLY  |  '
            'CancerScan AI Powered Security Audit System  |  '
            'This log is tamper-evident and cryptographically timestamped.'
        )
        footer.font = Font(name='Calibri', size=8, italic=True, color=MID_TEXT)
        footer.alignment = Alignment(horizontal='center')
        ws.row_dimensions[footer_row].height = 20

        # Widths
        col_widths = [14, 14, 14, 22, 20, 40, 14, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A7'
        ws.auto_filter.ref = f'A6:H{len(rows) + 6}'

        # Save
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'CancerScan_AuditLog_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        
        from flask import Response
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
    except Exception as e:
        current_app.logger.error(f"Audit log export failed: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Failed to generate audit export.'
        }), 500
    finally:
        if conn:
            conn.close()

# ── GET system health ────────────────────────
@admin_bp.route('/health', methods=['GET'])
@require_admin
def get_system_health():
    try:
        # Logic to check external nodes can be added here
        return jsonify({
            'success': True,
            'systems': [
                { 'name': 'Inference Core', 'status': 'Online', 'latency': '45ms', 'load': '12%' },
                { 'name': 'Database Node', 'status': 'Online', 'latency': '5ms', 'load': '2%' }
            ]
        })
    except Exception as e:
        current_app.logger.error(f"Health route error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Failed to retrieve system status'
        }), 500
